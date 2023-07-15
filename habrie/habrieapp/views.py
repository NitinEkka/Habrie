from datetime import date
from django.http import HttpResponse
from django.shortcuts import render
from .models import Student, Parent, AcademicDetail, Document
from .tasks import csv_upload, send_templated_email
from django.views.decorators.csrf import csrf_exempt
from .forms import StudentForm, ParentForm, AcademicDetailForm, DocumentForm
from django.utils.crypto import get_random_string
from datetime import datetime
from csvs.forms import CsvModelForm
from django.db.models import Q
import os
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa


def generate_enrollment_number(student_name):
    current_date = date.today().strftime('%d%m%y')
    initials = student_name[:3].upper()
    random_number = get_random_string(length=3, allowed_chars='0123456789')
    enrollment_number = current_date + initials + random_number
    return enrollment_number


def all_data(request):
    
    students = Student.objects.all()
    class_filter = request.GET.get('class_filter')
    section_filter = request.GET.get('section_filter')
    admission_cat_filter = request.GET.get('admission_cat_filter')
    classes = AcademicDetail.objects.values_list('class_id', flat=True).distinct()
    sections = AcademicDetail.objects.values_list('section_id', flat=True).distinct()
    admission_categories = Student.objects.values_list('addmission_cat', flat=True).distinct()

    students = Student.objects.all()
    if class_filter:
        students = students.filter(enroll_number__class_id=class_filter)
    if section_filter:
        students = students.filter(enroll_number__section_id=section_filter)
    if admission_cat_filter:
        students = students.filter(addmission_cat=admission_cat_filter)

    context = {
        'students': students,
        'classes': classes,
        'sections': sections,
        'admission_categories': admission_categories
    }

    return render(request, 'all_data.html', context)



@csrf_exempt
def create_student(request):
    context = {}
    upload_form = CsvModelForm(request.POST or None, request.FILES or None)
    
    if request.method == 'POST':
        student_form = StudentForm(request.POST)
        parent_form = ParentForm(request.POST)
        academic_form = AcademicDetailForm(request.POST)

        if student_form.is_valid() and parent_form.is_valid() and academic_form.is_valid():
            student = student_form.save(commit=False)
            student_name = student.name
            enrollment_number = generate_enrollment_number(student_name)
            academic = academic_form.save(commit=False)
            class_id = academic.class_id
            session_id = academic.session
            section_id = academic.section_id
            if enrollment_number:
                academic_detail = AcademicDetail.objects.create(enroll_id=enrollment_number, doj=datetime.now(), class_id=class_id, session=session_id, section_id=section_id)
                student.enroll_number = academic_detail
                student.save()

                parent = parent_form.save(commit=False)
                parent.student_name = student
                parent.save()

                academic = academic_form.save(commit=False)
                academic.student_name = student
                academic.save()

                empty_objects = AcademicDetail.objects.filter(enroll_id='')
                empty_objects.delete()

                recipient_email = student_form.cleaned_data.get('mail')
                recipient_name = student_form.cleaned_data.get('name')
                enroll_no = enrollment_number
                class_id = academic_form.cleaned_data.get('class_id')
                section = academic_form.cleaned_data.get('section_id')
                session = academic_form.cleaned_data.get('session')

                send_templated_email.delay(recipient_email,recipient_name, enroll_no, class_id, section, session)


                
        if upload_form.is_valid():
            upload_form.save()
            upload_form = CsvModelForm()
            csv_upload.delay()

            return HttpResponse("Single and bulk upload successfull")
        
        empty_objects = AcademicDetail.objects.filter(enroll_id='')
        empty_objects.delete()

        return HttpResponse("Single and bulk Upload Successful.")

    else:
        student_form = StudentForm()
        parent_form = ParentForm()
        academic_form = AcademicDetailForm()

    context = {
        'student_form': student_form,
        'parent_form': parent_form,
        'academic_form': academic_form,
        'upload_form': upload_form,
    }

    return render(request, 'index1.html', context)



















def export_to_pdf(request):
    class_filter = request.GET.get('class_filter')
    section_filter = request.GET.get('section_filter')
    admission_cat_filter = request.GET.get('admission_cat_filter')

    students = Student.objects.all()
    if class_filter:
        students = students.filter(enroll_number__class_id=class_filter)
    if section_filter:
        students = students.filter(enroll_number__section_id=section_filter)
    if admission_cat_filter:
        students = students.filter(addmission_cat=admission_cat_filter)

    template = get_template('pdf_template.html')
    context = {
        'students': students,
        'class_filter': class_filter,
        'section_filter': section_filter,
        'admission_cat_filter': admission_cat_filter,
    }
    html = template.render(context)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="student_list.pdf"'

    pisa.CreatePDF(html, dest=response)

    return response








from django.http import HttpResponse
from openpyxl import Workbook

def export_to_excel(request):
    
    class_filter = request.GET.get('class_filter')
    section_filter = request.GET.get('section_filter')
    admission_cat_filter = request.GET.get('admission_cat_filter')

    
    students = Student.objects.all()
    if class_filter:
        students = students.filter(enroll_number__class_id=class_filter)
    if section_filter:
        students = students.filter(enroll_number__section_id=section_filter)
    if admission_cat_filter:
        students = students.filter(addmission_cat=admission_cat_filter)

    
    workbook = Workbook()
    sheet = workbook.active

    
    column_names = ['Name', 'Gender', 'Aadhar', 'Date of Birth', 'Admission Category', 'Class', 'Section', 'Height', 'Weight', 'Mail', 'Contact', 'Address']
    for index, column_name in enumerate(column_names, start=1):
        sheet.cell(row=1, column=index, value=column_name)

    
    for row, student in enumerate(students, start=2):
        sheet.cell(row=row, column=1, value=student.name)
        sheet.cell(row=row, column=2, value=student.gender)
        sheet.cell(row=row, column=3, value=student.aadhar)
        sheet.cell(row=row, column=4, value=student.dob)
        sheet.cell(row=row, column=5, value=student.addmission_cat)
        sheet.cell(row=row, column=6, value=student.enroll_number.class_id)
        sheet.cell(row=row, column=7, value=student.enroll_number.section_id)
        sheet.cell(row=row, column=8, value=student.height)
        sheet.cell(row=row, column=9, value=student.weight)
        sheet.cell(row=row, column=10, value=student.mail)
        sheet.cell(row=row, column=11, value=student.contact)
        sheet.cell(row=row, column=12, value=student.address)

    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="filtered_student_list.xlsx"'

    
    workbook.save(response)

    return response




from django.shortcuts import render, redirect
from .forms import DocumentForm

def upload_document(request):
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('document_list')  
    else:
        form = DocumentForm()
    
    context = {'form': form}
    return render(request, 'document_upload.html', context)



   
